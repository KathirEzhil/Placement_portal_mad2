import os

from datetime import datetime

from openpyxl import Workbook


EXPORT_FOLDER = "exports"

os.makedirs(
    EXPORT_FOLDER,
    exist_ok=True
)


def generate_excel(headers, rows, filename_prefix):

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    filename = (
        f"{filename_prefix}_{timestamp}.xlsx"
    )

    filepath = os.path.join(
        EXPORT_FOLDER,
        filename
    )

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Export"

    worksheet.append(headers)

    for row in rows:

        worksheet.append(row)

    workbook.save(filepath)

    return filepath


def generate_multi_sheet_excel(sheets, filename_prefix):

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    filename = (
        f"{filename_prefix}_{timestamp}.xlsx"
    )

    filepath = os.path.join(
        EXPORT_FOLDER,
        filename
    )

    workbook = Workbook()

    # Remove the default worksheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, data in sheets.items():

        headers = data["headers"]
        rows = data["rows"]

        worksheet = workbook.create_sheet(
            title=sheet_name
        )

        # Add headers
        worksheet.append(headers)

        # Add data
        for row in rows:
            worksheet.append(row)

        # Freeze header row
        worksheet.freeze_panes = "A2"

        # Enable filters
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        # Make columns readable
        for column in worksheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                40
            )

    workbook.save(filepath)

    return filepath